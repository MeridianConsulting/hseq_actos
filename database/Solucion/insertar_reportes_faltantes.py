"""
Script para insertar en SQL los reportes del Excel que no están en el SQL.
Lee: Tablero Visibilidad Subsuelo 2026.xlsx y hseq_12_03_2026.sql
Genera: reportes_faltantes.sql con los INSERT statements.
"""

import re
import unicodedata
from datetime import datetime
import openpyxl

EXCEL_FILE = "Tablero Visibilidad Subsuelo 2026.xlsx"
SQL_FILE   = "hseq_12_03_2026.sql"
OUTPUT_FILE = "reportes_faltantes.sql"


# ---------------------------------------------------------------------------
# Utilidades
# ---------------------------------------------------------------------------

def normalizar(texto):
    """Convierte a minúsculas y elimina tildes para comparación flexible."""
    if not texto:
        return ""
    txt = str(texto).strip().lower()
    return "".join(
        c for c in unicodedata.normalize("NFD", txt)
        if unicodedata.category(c) != "Mn"
    )


def escapar_sql(valor):
    """Escapa comillas simples para SQL."""
    if valor is None:
        return "NULL"
    return "'" + str(valor).replace("\\", "\\\\").replace("'", "\\'") + "'"


def fecha_a_sql_date(valor):
    """Convierte un valor de fecha (objeto date, string) a 'YYYY-MM-DD'."""
    if valor is None:
        return "NULL"
    if hasattr(valor, "strftime"):
        return escapar_sql(valor.strftime("%Y-%m-%d"))
    s = str(valor).strip()
    # Intentar formato d/m/yyyy (español)
    for fmt in ("%d/%m/%Y", "%m/%d/%Y", "%Y-%m-%d"):
        try:
            return escapar_sql(datetime.strptime(s, fmt).strftime("%Y-%m-%d"))
        except ValueError:
            pass
    return escapar_sql(s)


def fecha_a_sql_timestamp(valor):
    """Convierte fecha a timestamp 'YYYY-MM-DD HH:MM:SS'."""
    if valor is None:
        return "NULL"
    if hasattr(valor, "strftime"):
        return escapar_sql(valor.strftime("%Y-%m-%d %H:%M:%S"))
    s = str(valor).strip()
    # Si ya tiene hora
    for fmt in ("%Y-%m-%d %H:%M:%S", "%d/%m/%Y %H:%M:%S"):
        try:
            return escapar_sql(datetime.strptime(s, fmt).strftime("%Y-%m-%d %H:%M:%S"))
        except ValueError:
            pass
    # Solo fecha
    for fmt in ("%d/%m/%Y", "%m/%d/%Y", "%Y-%m-%d"):
        try:
            return escapar_sql(datetime.strptime(s, fmt).strftime("%Y-%m-%d 00:00:00"))
        except ValueError:
            pass
    return escapar_sql(s + " 00:00:00")


# ---------------------------------------------------------------------------
# Paso 1: Leer SQL y extraer IDs existentes + mapa de usuarios
# ---------------------------------------------------------------------------

def extraer_ids_sql(ruta_sql):
    """Retorna set de IDs de reportes ya presentes en el SQL."""
    ids = set()
    with open(ruta_sql, encoding="utf-8", errors="replace") as f:
        contenido = f.read()
    patron = re.compile(
        r"INSERT INTO `reportes`.*?VALUES\s*(.*?);\s*\n", re.DOTALL
    )
    m = patron.search(contenido)
    if not m:
        return ids
    bloque = m.group(1)
    for m2 in re.finditer(r"\((\d+),", bloque):
        ids.add(int(m2.group(1)))
    return ids


def extraer_usuarios_sql(ruta_sql):
    """Retorna dict {nombre_normalizado: id} de la tabla usuarios."""
    usuarios = {}
    with open(ruta_sql, encoding="utf-8", errors="replace") as f:
        contenido = f.read()
    patron = re.compile(
        r"INSERT INTO `usuarios`.*?VALUES\s*(.*?);\s*\n", re.DOTALL
    )
    m = patron.search(contenido)
    if not m:
        return usuarios
    bloque = m.group(1)
    # Capturar (id, 'nombre', ...) -- el nombre es el 2do campo
    for m2 in re.finditer(r"\((\d+),\s*'([^']+)'", bloque):
        uid  = int(m2.group(1))
        nombre = m2.group(2)
        usuarios[normalizar(nombre)] = uid
    return usuarios


# ---------------------------------------------------------------------------
# Paso 2: Parsear Excel – formato vertical clave/valor
# ---------------------------------------------------------------------------

def parsear_hoja(ws, tipo_esperado):
    """
    Parsea una hoja del Excel con formato vertical:
      REPORTE N - TIPO
      ID  | N
      Tipo | ...
      ...
    Retorna lista de dicts con todos los campos del reporte.
    """
    reportes = []
    reporte_actual = None

    for fila in ws.iter_rows(values_only=True):
        clave = str(fila[0]).strip() if fila[0] is not None else ""
        valor = fila[1] if len(fila) > 1 else None

        # Encabezado de nuevo reporte
        if clave.startswith("REPORTE ") and ("-" in clave):
            if reporte_actual:
                reportes.append(reporte_actual)
            reporte_actual = {"_tipo_hoja": tipo_esperado}
            continue

        if reporte_actual is None:
            continue

        # Mapeo de claves del Excel a campos internos
        mapeo = {
            "id": "id",
            "tipo": "tipo_reporte",
            "estado": "estado",
            "usuario": "_nombre_usuario",
            "proyecto": "_proyecto",
            "asunto": "asunto",
            "fecha evento": "fecha_evento",
            "fecha creaci\u00f3n": "creado_en",
            "fecha creacion": "creado_en",
            "lugar hallazgo": "lugar_hallazgo",
            "tipo hallazgo": "tipo_hallazgo",
            "estado condici\u00f3n": "estado_condicion",
            "estado condicion": "estado_condicion",
            "descripci\u00f3n": "_descripcion",
            "descripcion": "_descripcion",
            "recomendaciones": "recomendaciones",
            "tipo conversaci\u00f3n": "tipo_conversacion",
            "tipo conversacion": "tipo_conversacion",
            "sitio evento": "sitio_evento_conversacion",
            "lugar conversaci\u00f3n": "lugar_hallazgo_conversacion",
            "lugar conversacion": "lugar_hallazgo_conversacion",
            "grado criticidad": "grado_criticidad",
        }

        clave_norm = normalizar(clave)
        campo = mapeo.get(clave_norm)
        if campo:
            reporte_actual[campo] = valor

    if reporte_actual:
        reportes.append(reporte_actual)

    return reportes


def parsear_excel(ruta_excel):
    """Retorna lista de todos los reportes encontrados en el Excel."""
    wb = openpyxl.load_workbook(ruta_excel)
    todos = []
    if "Hallazgos" in wb.sheetnames:
        todos += parsear_hoja(wb["Hallazgos"], "hallazgos")
    if "Conversaciones" in wb.sheetnames:
        todos += parsear_hoja(wb["Conversaciones"], "conversaciones")
    return todos


# ---------------------------------------------------------------------------
# Paso 3: Resolver id_usuario
# ---------------------------------------------------------------------------

def resolver_usuario(nombre_excel, mapa_usuarios):
    """
    Busca el id del usuario. Primero búsqueda exacta normalizada,
    luego por contiene.
    """
    if nombre_excel is None:
        return None
    clave = normalizar(str(nombre_excel))
    if clave in mapa_usuarios:
        return mapa_usuarios[clave]
    # Búsqueda parcial: si el nombre Excel está contenido en alguna clave del mapa
    for k, v in mapa_usuarios.items():
        if clave in k or k in clave:
            return v
    return None


# ---------------------------------------------------------------------------
# Paso 4: Generar INSERT SQL
# ---------------------------------------------------------------------------

COLUMNAS = (
    "id", "id_usuario", "telefono_contacto", "correo_contacto",
    "tipo_reporte", "tipo_pqr", "asunto", "descripcion_general",
    "fecha_evento", "lugar_hallazgo", "lugar_hallazgo_otro",
    "tipo_hallazgo", "descripcion_hallazgo", "recomendaciones",
    "estado_condicion", "grado_criticidad", "ubicacion_incidente",
    "hora_evento", "tipo_afectacion", "descripcion_incidente",
    "tipo_conversacion", "sitio_evento_conversacion",
    "lugar_hallazgo_conversacion", "lugar_hallazgo_conversacion_otro",
    "descripcion_conversacion", "asunto_conversacion",
    "estado", "revisado_por", "comentarios_revision", "fecha_revision",
    "creado_en", "actualizado_en",
)


def reporte_a_insert(r, mapa_usuarios):
    """Genera un INSERT INTO reportes ... VALUES (...); para un reporte."""
    tipo = r.get("tipo_reporte") or r.get("_tipo_hoja", "hallazgos")
    nombre_usuario = r.get("_nombre_usuario")
    id_usuario = resolver_usuario(nombre_usuario, mapa_usuarios)
    if id_usuario is None:
        id_usuario = "NULL  -- USUARIO NO ENCONTRADO: " + str(nombre_usuario)
        id_u_sql = f"NULL  /* USUARIO NO ENCONTRADO: {nombre_usuario} */"
    else:
        id_u_sql = str(id_usuario)

    desc = r.get("_descripcion")
    descripcion_hallazgo   = escapar_sql(desc) if tipo == "hallazgos"      else "NULL"
    descripcion_conversacion = escapar_sql(desc) if tipo == "conversaciones" else "NULL"

    asunto = r.get("asunto")
    asunto_conversacion = escapar_sql(asunto) if tipo == "conversaciones" else "NULL"

    creado_en      = fecha_a_sql_timestamp(r.get("creado_en"))
    actualizado_en = creado_en

    vals = {
        "id":                              str(int(r["id"])),
        "id_usuario":                      id_u_sql,
        "telefono_contacto":               "NULL",
        "correo_contacto":                 "NULL",
        "tipo_reporte":                    escapar_sql(tipo),
        "tipo_pqr":                        "NULL",
        "asunto":                          escapar_sql(asunto),
        "descripcion_general":             "NULL",
        "fecha_evento":                    fecha_a_sql_date(r.get("fecha_evento")),
        "lugar_hallazgo":                  escapar_sql(r.get("lugar_hallazgo")) if tipo == "hallazgos" else "NULL",
        "lugar_hallazgo_otro":             escapar_sql(r.get("lugar_hallazgo_otro", "")),
        "tipo_hallazgo":                   escapar_sql(r.get("tipo_hallazgo")) if tipo == "hallazgos" else "NULL",
        "descripcion_hallazgo":            descripcion_hallazgo,
        "recomendaciones":                 escapar_sql(r.get("recomendaciones")) if tipo == "hallazgos" else "NULL",
        "estado_condicion":                escapar_sql(r.get("estado_condicion")) if tipo == "hallazgos" else "NULL",
        "grado_criticidad":                escapar_sql(r.get("grado_criticidad")),
        "ubicacion_incidente":             "NULL",
        "hora_evento":                     "NULL",
        "tipo_afectacion":                 "NULL",
        "descripcion_incidente":           "NULL",
        "tipo_conversacion":               escapar_sql(r.get("tipo_conversacion")) if tipo == "conversaciones" else "NULL",
        "sitio_evento_conversacion":       escapar_sql(r.get("sitio_evento_conversacion")) if tipo == "conversaciones" else "NULL",
        "lugar_hallazgo_conversacion":     escapar_sql(r.get("lugar_hallazgo_conversacion")) if tipo == "conversaciones" else "NULL",
        "lugar_hallazgo_conversacion_otro": "''",
        "descripcion_conversacion":        descripcion_conversacion,
        "asunto_conversacion":             asunto_conversacion,
        "estado":                          escapar_sql(r.get("estado", "pendiente")),
        "revisado_por":                    "NULL",
        "comentarios_revision":            "NULL",
        "fecha_revision":                  "NULL",
        "creado_en":                       creado_en,
        "actualizado_en":                  actualizado_en,
    }

    cols = ", ".join(f"`{c}`" for c in COLUMNAS)
    vlist = ", ".join(vals[c] for c in COLUMNAS)
    return f"INSERT INTO `reportes` ({cols}) VALUES\n({vlist});"


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    print("Leyendo SQL existente...")
    ids_sql      = extraer_ids_sql(SQL_FILE)
    mapa_usuarios = extraer_usuarios_sql(SQL_FILE)
    print(f"  IDs en SQL: {sorted(ids_sql)}")
    print(f"  Usuarios encontrados: {len(mapa_usuarios)}")

    print("\nLeyendo Excel...")
    reportes_excel = parsear_excel(EXCEL_FILE)
    print(f"  Reportes en Excel: {len(reportes_excel)}")

    # Filtrar los que NO están en el SQL
    faltantes = [
        r for r in reportes_excel
        if r.get("id") is not None and int(r["id"]) not in ids_sql
    ]
    # Ordenar por ID ascendente
    faltantes.sort(key=lambda r: int(r["id"]))

    print(f"\n  Reportes faltantes ({len(faltantes)}): {[int(r['id']) for r in faltantes]}")

    # Generar SQL de salida
    lineas = [
        "-- ============================================================",
        f"-- Reportes del Excel que NO están en hseq_12_03_2026.sql",
        f"-- Generado: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
        f"-- Total de inserciones: {len(faltantes)}",
        "-- ============================================================",
        "",
        "SET SQL_MODE = 'NO_AUTO_VALUE_ON_ZERO';",
        "START TRANSACTION;",
        "SET NAMES utf8mb4;",
        "",
    ]

    no_encontrados = []
    for r in faltantes:
        try:
            nombre = r.get("_nombre_usuario")
            uid = resolver_usuario(nombre, mapa_usuarios)
            if uid is None:
                no_encontrados.append((int(r["id"]), nombre))
            lineas.append(f"-- Reporte {int(r['id'])} - {r.get('tipo_reporte', r.get('_tipo_hoja'))} ({nombre})")
            lineas.append(reporte_a_insert(r, mapa_usuarios))
            lineas.append("")
        except Exception as e:
            lineas.append(f"-- ERROR al procesar reporte {r.get('id')}: {e}")
            lineas.append("")

    lineas.append("COMMIT;")
    lineas.append("")

    with open(OUTPUT_FILE, "w", encoding="utf-8") as f:
        f.write("\n".join(lineas))

    print(f"\nArchivo generado: {OUTPUT_FILE}")

    if no_encontrados:
        print("\n[!] USUARIOS NO ENCONTRADOS (revisar manualmente):")
        for rid, nombre in no_encontrados:
            print(f"  - Reporte {rid}: '{nombre}'")
    else:
        print("[OK] Todos los usuarios fueron identificados correctamente.")


if __name__ == "__main__":
    main()
